from fastapi import FastAPI, File, Query, UploadFile, HTTPException
from fastapi.responses import JSONResponse
import httpx
import openpyxl
import os
import io
import pandas as pd
from docxtpl import DocxTemplate
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

app = FastAPI()

# Variables de configuration
BASE_DIR = os.getcwd()
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "template", "S1")
TEMPLATE_FILE = os.path.join(BASE_DIR, "template", "modeleM1S1.docx")

base_url = os.getenv('YPAERO_BASE_URL')
jeton = os.getenv('YPAERO_API_TOKEN')


def format_duration_to_minutes(duration_str):
    """ Converts a string duration into total minutes. """
    parts = duration_str.split('h')
    if len(parts) == 2:
        hours = int(parts[0])
        minutes = int(parts[1])
        return hours * 60 + minutes
    return int(duration_str.split()[0])  # Assuming format is "XX minutes"

def format_minutes_to_duration(minutes):
    """ Formats minutes into hour and minute format. """
    if minutes == 0:
        return "00h00"
    hours = minutes // 60
    remaining_minutes = minutes % 60
    if hours > 0:
        return f"{hours}h{remaining_minutes:02d}"
    return f"{remaining_minutes} minutes"

def sum_durations(duration_list):
    """ Sums a list of durations given in mixed formats. """
    total_minutes = sum(format_duration_to_minutes(d) for d in duration_list)
    return format_minutes_to_duration(total_minutes)

def fetch_api_data(url, headers):
    with httpx.Client() as client:
        response = client.get(url, headers=headers, timeout=60.0)
        if response.status_code == 200:
            return response.json()
        else:
            return None

def extract_grades_and_coefficients(grade_str):
    grades_coefficients = []
    parts = grade_str.split(" - ")
    for part in parts:
        # Vérifie si la part indique une absence au devoir, avec ou sans coefficient spécifié
        if "Absent au devoir" in part:
            continue  # Ignore cette part et passe à la suivante sans ajouter à grades_coefficients

        if "(" in part:
            grade_part, coefficient_part = part.split("(")
            coefficient = coefficient_part.replace(")", "").replace(",", ".")  # Convertir la virgule en point pour le coefficient
            grade = grade_part.replace(",", ".")  # Convertir les virgules en points pour les notes
        else:
            grade = part.replace(",", ".")  # Convertir les virgules en points pour les notes sans coefficient explicite
            coefficient = "1.0"  # Coefficient implicite de 1, comme décimal
        grade = grade.strip()
        coefficient = coefficient.strip()
        # Debugging print pour vérifier la conversion
        # print(f"Traitement de la part: '{part}', grade converti: '{grade}', coefficient converti: '{coefficient}'")
        grades_coefficients.append((float(grade), float(coefficient)))
    return grades_coefficients

def calculate_weighted_average_from_string(grade_str):
    grades_with_coefficients = extract_grades_and_coefficients(grade_str)
    total_grade = 0.0
    total_coefficient = 0.0
    for grade, coefficient in grades_with_coefficients:
        # Ajoutons un print pour déboguer
        # print(f"Multiplication de la note {grade} par son coefficient {coefficient}")
        total_grade += grade * coefficient
        total_coefficient += coefficient
    return total_grade / total_coefficient if total_coefficient else 0


@app.get("/get-apprenants/")
async def get_apprenants():
    endpoint_apprenant = "/r/v1/formation-longue/apprenants?codesPeriode=2"
    url_apprenant = f"{base_url}{endpoint_apprenant}"
    headers = {
        "X-Auth-Token": jeton,
        "Content-Type": "application/json"
    }
    async with httpx.AsyncClient() as client:
        response = await client.get(url_apprenant, headers=headers, timeout=30)
        
        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(status_code=response.status_code, detail="Erreur lors de l'appel API pour les apprenants")

@app.get("/get-groupes/")
async def get_groupes():
    endpoint = "/r/v1/formation-longue/groupes"
    url = f"{base_url}{endpoint}"
    headers = {
        "X-Auth-Token": jeton,
        "Content-Type": "application/json"
    }
    async with httpx.AsyncClient() as client:
        print(jeton)
        response = await client.get(url, headers=headers)
        
        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(status_code=response.status_code, detail="Erreur lors de l'appel API")
        
@app.get("/get-absences/")
async def get_absences(date_deb: str = Query(..., regex="^\\d{2}-\\d{2}-\\d{4}$"), date_fin: str = Query(..., regex="^\\d{2}-\\d{2}-\\d{4}$")):
    endpoint = f"/r/v1/absences/{date_deb}/{date_fin}"
    url = f"{base_url}{endpoint}"
    headers = {
        "X-Auth-Token": jeton,
        "Content-Type": "application/json"
    }
    async with httpx.AsyncClient() as client:
        response = await client.get(url, headers=headers)
        
        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(status_code=response.status_code, detail="Erreur lors de l'appel API pour les absences")


@app.post('/upload')
async def upload_file(file: UploadFile = File(...)):
    try:
        # Save uploaded file to disk first
        temp_file_path = os.path.join(UPLOAD_DIR, file.filename)
        with open(temp_file_path, 'wb') as temp_file:
            content = await file.read()  # Read file content asynchronously
            temp_file.write(content)
        
        # Now open the file with openpyxl
        wb = openpyxl.load_workbook(temp_file_path)
        ws = wb.active

        # Phrase à supprimer
        target_phrase = "* Attention, le total des absences prend en compte toutes les absences aux séances sur la période concernée. S'il existe des absences sur des matières qui ne figurent pas dans le relevé, elles seront également comptabilisées."
        #Supprimer la phrase
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == target_phrase:
                    cell.value = None

        # Sauvegarder le workbook modifié dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Définition des en-têtes pour le fichier Excel
        headers = ['A5', 'B5', 'BN5', 'BO5', 'BP5', 'BQ5', 'BR5', 'BS5', 'BT5', 'BU5']
        titles = ["Code Apprenant", "Nom", "Date de Naissance", "Nom Site", "Code Groupe", "Nom Groupe", "Étendu Groupe", "ABS justifiées", "ABS injustifiées", "Retards"]
        for header, title in zip(headers, titles):
            ws[header] = title

        apprenants_data = fetch_api_data(f"{base_url}/r/v1/formation-longue/apprenants?codesPeriode=2", {"X-Auth-Token": jeton, "Content-Type": "application/json"})
        groupes_data = fetch_api_data(f"{base_url}/r/v1/formation-longue/groupes", {"X-Auth-Token": jeton, "Content-Type": "application/json"})
        absences_data = fetch_api_data(f"{base_url}/r/v1/absences/01-01-2023/31-12-2024", {"X-Auth-Token": jeton, "Content-Type": "application/json"})

        if not apprenants_data or not groupes_data:
            # return jsonify({"error": "Impossible de récupérer les données des apprenants"}), 500
            return JSONResponse(content={"message": "Impossible de récupérer les données des apprenants"})

        # Création de dictionnaires pour les apprenants et les groupes
        apprenants_dict = {
            f"{app['nomApprenant'].upper()} {app['prenomApprenant'].upper()}": app
            for app in apprenants_data.values()
        }
        
        # Création de dictionnaires pour les groupes
        groupes_dict = {
            str(groupe['codeGroupe']): (groupe['nomGroupe'], groupe.get('etenduGroupe', 'N/A'))
            for groupe in groupes_data.values()
        }

        absences_summary = {}
        for absence in absences_data.values():
            apprenant_id = absence.get('codeApprenant')
            duration = int(absence.get('duree', 0))
            formatted_duration = format_minutes_to_duration(duration)

            if apprenant_id not in absences_summary:
                absences_summary[apprenant_id] = {'justified': [], 'unjustified': [], 'delays': []}

            if absence.get('isJustifie'):
                absences_summary[apprenant_id]['justified'].append(formatted_duration)
            elif absence.get('isRetard'):
                absences_summary[apprenant_id]['delays'].append(formatted_duration)
            else:
                absences_summary[apprenant_id]['unjustified'].append(formatted_duration)


        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=2):
            cell_nom = row[0]
            nom_complet = cell_nom.value.upper() if cell_nom.value else ''
            
            apprenant_correspondant = apprenants_dict.get(nom_complet)
            if apprenant_correspondant:
                ws[f'A{cell_nom.row}'] = apprenant_correspondant.get('codeApprenant', 'N/A')
                ws[f'BN{cell_nom.row}'] = apprenant_correspondant.get('dateNaissance', 'N/A')
                # Naviguer à travers les données JSON pour extraire nomSite
                nom_site = 'N/A'
                for inscription in apprenant_correspondant.get('inscriptions', []):
                    site_info = inscription.get('site', {})
                    if site_info.get('nomSite'):
                        nom_site = site_info['nomSite']
                        break  # Prendre le premier site valide
                ws[f'BO{cell_nom.row}'] = nom_site

                # Extraction du codeGroupe depuis informationsCourantes
                code_groupe = apprenant_correspondant.get('informationsCourantes', {}).get('codeGroupe', 'N/A')
                ws[f'BP{cell_nom.row}'] = code_groupe
                groupe_info = groupes_dict.get(str(code_groupe), ('N/A', 'N/A'))
                ws[f'BQ{cell_nom.row}'] = groupe_info[0]
                ws[f'BR{cell_nom.row}'] = groupe_info[1] 

                # Extraction des absences
                apprenant_id = apprenant_correspondant.get('codeApprenant')
                abs_info = absences_summary.get(apprenant_id, {'justified': [], 'unjustified': [], 'delays': []})
                ws[f'BS{cell_nom.row}'] = sum_durations(abs_info['justified']) or "00h00"
                ws[f'BT{cell_nom.row}'] = sum_durations(abs_info['unjustified']) or "00h00"
                ws[f'BU{cell_nom.row}'] = sum_durations(abs_info['delays']) or "00h00"

        # Sauvegardez le workbook modifié
        modified_file_path = os.path.join(UPLOAD_DIR, 'modified_file.xlsx')
        wb.save(modified_file_path)

        # Supprimez le fichier temporaire si nécessaire
        os.remove(temp_file_path)

        return {"message": "Excel file processed successfully", "filename": modified_file_path}


    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def process_excel_file(file_path: str, template_path: str, output_dir: str) -> list:
    try:
        # Load the Excel file, skipping headers for the main data as before
        df_titles = pd.read_excel(file_path, header=None)
        titles_row = df_titles.iloc[3]  # Cette ligne peut ne pas être nécessaire si non utilisée
        
        # Reload for student data, assuming headers start from row 4 (index 3 in Python)
        df_students = pd.read_excel(file_path, header=4)
        
        bulletin_paths = []

        # Les positions des colonnes ECTS dans le fichier Excel, en assumant que la première colonne est indexée à 0
        ects_columns = [
            3, 15, 21, 30, 51,  # Indices pour D6, P6, V6, AE6, AZ6 (ECTSUE1 à ECTSUE5)
            6, 9, 12, 18, 24,   # Indices pour G6, J6, M6, S6, Y6
            27, 33, 36, 39,     # Indices pour AB6, AH6, AK6, AN6
            42, 48,             # Indices pour AQ6, AW6
            45, 54, 57, 60      # Indices pour AT6, BC6, BF6, BI6
        ]

        # Iterate over each student in the DataFrame
        for index, student_data in df_students.iterrows():
            # Call the document generation function for the current student
            ects_values = [student_data.iloc[col] for col in ects_columns]  # Utilisez iloc pour accéder par indice de colonne
            bulletin_path = generate_word_document(student_data, ects_values, titles_row, template_path, output_dir)
            bulletin_paths.append(bulletin_path)

        return bulletin_paths
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {e}")


def generate_word_document(student_data, ects_values, titles_row, template_path, output_dir):
    ects_values_as_int = [int(value) if pd.notna(value) else "NaN" for value in ects_values] 
    
    # Preparing the data for the placeholders
    placeholders = {
        "nomApprenant": student_data["Nom"],
        "etendugroupe": student_data["Étendu Groupe"],
        "dateNaissance": student_data["Date de Naissance"],
        "groupe": student_data["Nom Groupe"],
        "campus": student_data["Nom Site"],
        "justifiee": student_data["ABS justifiées"],
        "injustifiee": student_data["ABS injustifiées"],
        "retard": student_data["Retards"],
        "UE1_Title": titles_row[2],
        "strat": titles_row[5],
        "finance": titles_row[8],
        "economie": titles_row[11],
        "UE2_Title": titles_row[14],
        "droit": titles_row[17],
        "UE3_Title": titles_row[20],
        "ville": titles_row[23],
        "politique": titles_row[26],
        "UE4_Title": titles_row[29],
        "real": titles_row[32],
        "rencontre": titles_row[35],
        "career": titles_row[38],
        "inside": titles_row[41],
        "immersion": titles_row[44],
        "voltaire": titles_row[47],
        "UESPE_Title": titles_row[50],
        "fonciere": titles_row[53],
        "montage": titles_row[56],
        "acquisition": titles_row[59],
        # ECTS
        # Ajout des placeholders pour les ECTS
        "ECTSUE1": ects_values_as_int[0],
        "ECTSUE2": ects_values_as_int[1],
        "ECTSUE3": ects_values_as_int[2],
        "ECTSUE4": ects_values_as_int[3],
        "ECTSUE5": ects_values_as_int[4],
        "ECTS1": ects_values_as_int[5],
        "ECTS2": ects_values_as_int[6],
        "ECTS3": ects_values_as_int[7],
        "ECTS4": ects_values_as_int[8],
        "ECTS5": ects_values_as_int[9],
        "ECTS6": ects_values_as_int[10],
        "ECTS7": ects_values_as_int[11],
        "ECTS8": ects_values_as_int[16],
        "ECTS9": ects_values_as_int[17],
        "ECTS10": ects_values_as_int[18],
        "ECTS11": ects_values_as_int[19],
    }

    all_grades_with_coefficients = []  # Pour calculer la moyenne générale à la fin
    grade_column_indices = [5, 8, 11, 17, 23, 26, 32, 35, 38, 41, 44, 47, 53, 56, 59]  # Assurez-vous que ces indices correspondent

    for i, col_index in enumerate(grade_column_indices, start=1):
        grade_str = str(student_data.iat[col_index]).strip() if pd.notna(student_data.iat[col_index]) else ""
        if grade_str:
            # Assurez-vous que cette fonction renvoie une liste de tuples (note, coefficient)
            grades_coefficients = extract_grades_and_coefficients(grade_str)
            all_grades_with_coefficients.extend(grades_coefficients)  # Ajoutez les tuples à la liste générale

            # Calculer et afficher la moyenne individuelle
            individual_average = sum(grade * coefficient for grade, coefficient in grades_coefficients) / sum(coefficient for grade, coefficient in grades_coefficients) if grades_coefficients else 0
            placeholders[f"note{i}"] = f"{individual_average:.2f}"

    # Calcul de la moyenne générale pour toutes les notes et coefficients collectés
    if all_grades_with_coefficients:
        general_average = sum(grade * coefficient for grade, coefficient in all_grades_with_coefficients) / sum(coefficient for grade, coefficient in all_grades_with_coefficients)
        placeholders["moyenne"] = f"{general_average:.2f}"
    else:
        placeholders["moyenne"] = ""

    # Génération du document avec les placeholders remplis
    doc = DocxTemplate(template_path)
    # Génération du document avec les placeholders remplis
    doc.render(placeholders)
    
    output_filename = f"{student_data['Nom']}_bulletin.docx"
    output_filepath = os.path.join(output_dir, output_filename)
    doc.save(output_filepath)
    return output_filepath


@app.post("/upload-excel/")
async def upload_excel(file: UploadFile = File(...)):
    try:
        file_location = os.path.join(UPLOAD_DIR, file.filename)
        with open(file_location, "wb+") as file_object:
            file_object.write(await file.read())
        bulletin_paths = process_excel_file(file_location, TEMPLATE_FILE, OUTPUT_DIR)
        return {"filenames": bulletin_paths}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)