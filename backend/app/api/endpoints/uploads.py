import logging
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
import os
import openpyxl
import io
from app.core.config import settings
from app.services.excel_service import process_excel_file
from app.utils.date_utils import format_duration_to_minutes, format_minutes_to_duration, sum_durations
from app.services.api_service import fetch_api_data

# Configurer le logger
logger = logging.getLogger(__name__)

router = APIRouter()

@router.post("/upload-file")
async def upload_file(file: UploadFile = File(...)):
    try:
        # Save uploaded file to disk first
        temp_file_path = os.path.join(settings.UPLOAD_DIR, file.filename)
        with open(temp_file_path, 'wb') as temp_file:
            content = await file.read()  # Read file content asynchronously
            temp_file.write(content)
        
        # Now open the file with openpyxl
        wb = openpyxl.load_workbook(temp_file_path)
        ws = wb.active

        # Phrase à supprimer
        target_phrase = "* Attention, le total des absences prend en compte toutes les absences aux séances sur la période concernée. S'il existe des absences sur des matières qui ne figurent pas dans le relevé, elles seront également comptabilisées."
        #Supprimer la phrase
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == target_phrase:
                    cell.value = None

        # Sauvegarder le workbook modifié dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Définition des en-têtes pour le fichier Excel
        headers = ['A5', 'B5', 'BN5', 'BO5', 'BP5', 'BQ5', 'BR5', 'BS5', 'BT5', 'BU5']
        titles = ["Code Apprenant", "Nom", "Date de Naissance", "Nom Site", "Code Groupe", "Nom Groupe", "Étendu Groupe", "ABS justifiées", "ABS injustifiées", "Retards"]
        for header, title in zip(headers, titles):
            ws[header] = title

        apprenants_data = await fetch_api_data(f"{settings.YPAERO_BASE_URL}/r/v1/formation-longue/apprenants?codesPeriode=2", {"X-Auth-Token": settings.YPAERO_API_TOKEN, "Content-Type": "application/json"})
        groupes_data = await fetch_api_data(f"{settings.YPAERO_BASE_URL}/r/v1/formation-longue/groupes", {"X-Auth-Token": settings.YPAERO_API_TOKEN, "Content-Type": "application/json"})
        absences_data = await fetch_api_data(f"{settings.YPAERO_BASE_URL}/r/v1/absences/01-01-2023/31-12-2024", {"X-Auth-Token": settings.YPAERO_API_TOKEN, "Content-Type": "application/json"})

        if not apprenants_data or not groupes_data:
            return JSONResponse(content={"message": "Impossible de récupérer les données des apprenants"})

        # Création de dictionnaires pour les apprenants et les groupes
        apprenants_dict = {
            f"{app['nomApprenant'].upper()} {app['prenomApprenant'].upper()}": app
            for app in apprenants_data.values()
        }
        
        # Création de dictionnaires pour les groupes
        groupes_dict = {
            str(groupe['codeGroupe']): (groupe['nomGroupe'], groupe.get('etenduGroupe', 'N/A'))
            for groupe in groupes_data.values()
        }

        absences_summary = {}
        for absence in absences_data.values():
            apprenant_id = absence.get('codeApprenant')
            duration = int(absence.get('duree', 0))
            formatted_duration = format_minutes_to_duration(duration)

            if apprenant_id not in absences_summary:
                absences_summary[apprenant_id] = {'justified': [], 'unjustified': [], 'delays': []}

            if absence.get('isJustifie'):
                absences_summary[apprenant_id]['justified'].append(formatted_duration)
            elif absence.get('isRetard'):
                absences_summary[apprenant_id]['delays'].append(formatted_duration)
            else:
                absences_summary[apprenant_id]['unjustified'].append(formatted_duration)


        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=2):
            cell_nom = row[0]
            nom_complet = cell_nom.value.upper() if cell_nom.value else ''
            
            apprenant_correspondant = apprenants_dict.get(nom_complet)
            if apprenant_correspondant:
                ws[f'A{cell_nom.row}'] = apprenant_correspondant.get('codeApprenant', 'N/A')
                ws[f'BN{cell_nom.row}'] = apprenant_correspondant.get('dateNaissance', 'N/A')
                # Naviguer à travers les données JSON pour extraire nomSite
                nom_site = 'N/A'
                for inscription in apprenant_correspondant.get('inscriptions', []):
                    site_info = inscription.get('site', {})
                    if site_info.get('nomSite'):
                        nom_site = site_info['nomSite']
                        break  # Prendre le premier site valide
                ws[f'BO{cell_nom.row}'] = nom_site

                # Extraction du codeGroupe depuis informationsCourantes
                code_groupe = apprenant_correspondant.get('informationsCourantes', {}).get('codeGroupe', 'N/A')
                ws[f'BP{cell_nom.row}'] = code_groupe
                groupe_info = groupes_dict.get(str(code_groupe), ('N/A', 'N/A'))
                ws[f'BQ{cell_nom.row}'] = groupe_info[0]
                ws[f'BR{cell_nom.row}'] = groupe_info[1] 

                # Extraction des absences
                apprenant_id = apprenant_correspondant.get('codeApprenant')
                abs_info = absences_summary.get(apprenant_id, {'justified': [], 'unjustified': [], 'delays': []})
                ws[f'BS{cell_nom.row}'] = sum_durations(abs_info['justified']) or "00h00"
                ws[f'BT{cell_nom.row}'] = sum_durations(abs_info['unjustified']) or "00h00"
                ws[f'BU{cell_nom.row}'] = sum_durations(abs_info['delays']) or "00h00"

        # Sauvegardez le workbook modifié
        modified_file_path = os.path.join(settings.UPLOAD_DIR, 'modified_file.xlsx')
        wb.save(modified_file_path)

        # Supprimez le fichier temporaire si nécessaire
        os.remove(temp_file_path)

        return {"message": "Excel file processed successfully", "filename": modified_file_path}


    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    logger.debug("Début de traitement du fichier Excel pour bulletin.")
    try:
        file_location = os.path.join(settings.UPLOAD_DIR, file.filename)
        with open(file_location, "wb+") as file_object:
            file_object.write(await file.read())
            logger.debug("Fichier Excel écrit sur le disque.")
        bulletin_paths = process_excel_file(file_location, settings.TEMPLATE_FILE, settings.OUTPUT_DIR)
        logger.debug("Bulletin généré avec succès.")
        return {"filenames": bulletin_paths}
    except Exception as e:
        logger.exception("Erreur lors du traitement de l'upload Excel: %s", e)
        raise HTTPException(status_code=400, detail=str(e))
